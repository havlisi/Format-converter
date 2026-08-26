from openpyxl import Workbook, load_workbook
from core.xlsx_io import read_xlsx, write_xlsx


def test_write_then_read_round_trip(tmp_path):
    p = tmp_path / "out.xlsx"
    blocks = [("table", [["Name", "Age"], ["Bob", "30"]])]

    write_xlsx(blocks, str(p))
    result = read_xlsx(str(p))

    assert result == [("table", [["Name", "Age"], ["Bob", "30"]])]


def test_read_skips_fully_empty_rows(tmp_path):
    p = tmp_path / "in.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["a", "b"])
    ws.append([None, None])
    ws.append(["c", "d"])
    wb.save(p)

    result = read_xlsx(str(p))

    assert result == [("table", [["a", "b"], ["c", "d"]])]


def test_write_text_block_goes_in_column_a(tmp_path):
    p = tmp_path / "out2.xlsx"
    blocks = [("text", "hello"), ("table", [["x"]])]

    write_xlsx(blocks, str(p))
    result = read_xlsx(str(p))

    assert result == [("table", [["hello"], ["x"]])]


def test_write_long_multiline_text_block_splits_into_one_row_per_line(tmp_path):
    # A raw-text fallback block (the last resort when a PDF can't be reconstructed
    # into a table) can run to tens of thousands of characters across a long
    # document. Writing it into a single cell hits Excel's ~32,767-character cell
    # limit — openpyxl silently truncates on write, with no error — and even
    # short of that limit, one giant unreadable cell defeats the point of a
    # spreadsheet. Splitting on newlines into one row per line keeps every line
    # comfortably under the per-cell cap and makes the content actually readable.
    lines = [f"line {i} " + ("x" * 40) for i in range(1000)]
    content = "\n".join(lines)
    p = tmp_path / "out3.xlsx"
    blocks = [("text", content)]

    write_xlsx(blocks, str(p))
    result = read_xlsx(str(p))

    assert result == [("table", [[line] for line in lines])]


def test_write_xlsx_does_not_turn_formula_looking_strings_into_formulas(tmp_path):
    p = tmp_path / "formulas.xlsx"
    blocks = [("table", [["=SUM(A1:A2)", "+1+2", "-1-2", "@SUM(A1)", "plain"]])]

    write_xlsx(blocks, str(p))

    # Verify at the openpyxl level that no cell was stored as an actual formula.
    wb = load_workbook(str(p))
    ws = wb.active
    for cell in ws[1]:
        assert cell.data_type != "f"

    result = read_xlsx(str(p))
    assert result == [("table", [["=SUM(A1:A2)", "+1+2", "-1-2", "@SUM(A1)", "plain"]])]


def test_read_xlsx_falls_back_to_formula_text_when_no_cached_value(tmp_path):
    p = tmp_path / "uncached_formula.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Label", "Value"])
    ws.append(["total", "=SUM(A1:A1)"])  # written as a real formula, no cached result
    wb.save(p)

    result = read_xlsx(str(p))

    assert result == [("table", [["Label", "Value"], ["total", "=SUM(A1:A1)"]])]
