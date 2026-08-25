from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from core.types import Block
from typing import List

_FORMULA_PREFIXES = ("=", "+", "-", "@")
_WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
_MIN_COLUMN_WIDTH = 8
_MAX_COLUMN_WIDTH = 60


def read_xlsx(path: str) -> List[Block]:
    wb = load_workbook(path, data_only=True)
    wb_formulas = load_workbook(path, data_only=False)
    blocks: List[Block] = []
    for ws, ws_formulas in zip(wb.worksheets, wb_formulas.worksheets):
        rows = []
        for row, row_formulas in zip(ws.iter_rows(), ws_formulas.iter_rows()):
            str_row = []
            for cell, cell_formula in zip(row, row_formulas):
                value = cell.value
                if value is None and isinstance(cell_formula.value, str) and cell_formula.value.startswith("="):
                    # data_only=True gave us None because the workbook was never opened in a
                    # real spreadsheet app (so there's no cached formula result). Fall back to
                    # showing the formula text itself rather than silently dropping the cell.
                    value = cell_formula.value
                str_row.append("" if value is None else str(value))
            if any(cell != "" for cell in str_row):
                rows.append(str_row)
        if rows:
            blocks.append(("table", rows))
    return blocks


def _set_cell(ws, row_idx: int, col_idx: int, value) -> None:
    cell = ws.cell(row=row_idx, column=col_idx, value=value)
    if isinstance(value, str) and value[:1] in _FORMULA_PREFIXES:
        # Force literal-string interpretation so openpyxl (and Excel) doesn't treat
        # e.g. "=SUM(...)" or "-5-3" as a formula to evaluate — it's just text data.
        cell.data_type = "s"
    if isinstance(value, str) and "\n" in value:
        # Without this, Excel collapses an embedded newline into the same visual line
        # instead of actually breaking the cell — multi-line content becomes unreadable.
        cell.alignment = _WRAP_ALIGNMENT


def _autofit_columns(ws) -> None:
    # A cell that wraps (embedded newline) shouldn't stretch its column to its full
    # length — that column gets capped and relies on wrapping + row height instead,
    # so one long "Description"-style column doesn't force every column that wide.
    #
    # Single-cell "banner" rows (free-standing text blocks, e.g. a page's preamble
    # sharing column A with a real "Date" column below it) are skipped when sizing —
    # they're not representative of that column's actual data, and would otherwise
    # drag a narrow column wide just because one unrelated row happened to be long.
    widths: dict = {}
    for row in ws.iter_rows():
        nonempty = [c for c in row if c.value not in (None, "")]
        if len(nonempty) <= 1:
            continue
        for cell in nonempty:
            text = str(cell.value)
            longest_line = max((len(line) for line in text.split("\n")), default=0)
            cap = _MAX_COLUMN_WIDTH if "\n" in text else None
            width = min(longest_line, cap) if cap else longest_line
            widths[cell.column] = max(widths.get(cell.column, _MIN_COLUMN_WIDTH), width)
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = min(width + 2, _MAX_COLUMN_WIDTH)


def write_xlsx(blocks: List[Block], path: str) -> None:
    wb = Workbook()
    ws = wb.active
    row_idx = 1
    for kind, content in blocks:
        if kind == "text":
            _set_cell(ws, row_idx, 1, content)
            row_idx += 1
        else:
            for row in content:
                for col_idx, val in enumerate(row, start=1):
                    _set_cell(ws, row_idx, col_idx, val)
                row_idx += 1
    _autofit_columns(ws)
    wb.save(path)
